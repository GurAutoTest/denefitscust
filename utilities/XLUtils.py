import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)

# def writeData(file, sheetName, data):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     print("fffffffffff")
#     # if isinstance(data, list) and all(isinstance(d, dict) for d in data):
#     print("hhhhhhhhhhhhh")
#     headers = list(data[0].__dict__.keys())
#     sheet.append(headers)



#     for obj in data:
#        data = list(obj.__dict__.values())
#        sheet.append(data)
#        print(data)

#        print("insideeeeeeeeeeeeeee")
#            # sheet.cell(row=rownum, column=columnno).value = data
#        workbook.save(file)   